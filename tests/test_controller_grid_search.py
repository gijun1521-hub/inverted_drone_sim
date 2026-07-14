import hashlib
import json
import math
import tempfile
import unittest
from dataclasses import replace
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from analysis import controller_grid_search as controller_grid_search_module
from analysis.controller_grid_search import (
    CACHE_SCHEMA_VERSION,
    SCORE_SPECS,
    SHEET_NAMES,
    WORKFLOW_IMPLEMENTATION_FILES,
    Candidate,
    SearchScenario,
    ScenarioResultStore,
    WorkflowOptions,
    attitude_p_candidates,
    attitude_p_scenarios,
    best_parameter_rows,
    build_workflow_fingerprint,
    compute_scenario_metrics,
    direction_reversal_count,
    loiter_xy_candidates,
    loiter_xy_scenarios,
    moving_mass_gain_candidates,
    normalized_score,
    path_length,
    rate_i_candidates,
    rate_i_scenarios,
    rate_pd_coarse_candidates,
    rate_pd_scenarios,
    require_best_aggregate_row,
    resolve_profile_output_directory,
    run_workflow,
    selected_parameter_boundary_rows,
    stage_validity_reasons,
    top_candidates,
    validate_previous_stage_metadata,
    write_profiles,
    zero_crossing_count,
)
from analysis.headless_loiter import LoiterRunResult, LoiterScenarioConfig
from params import load_interactive_config


class ControllerGridDefinitionTests(unittest.TestCase):
    def test_full_grid_endpoints_and_counts(self):
        rate_pd = rate_pd_coarse_candidates()
        self.assertEqual(len(rate_pd), 16 * 17)
        self.assertEqual(rate_pd[0].parameters, {"atc_rat_pit_p": 0.005, "atc_rat_pit_i": 0.0, "atc_rat_pit_d": 0.0})
        self.assertEqual(rate_pd[-1].parameters, {"atc_rat_pit_p": 0.08, "atc_rat_pit_i": 0.0, "atc_rat_pit_d": 0.008})

        rate_i = rate_i_candidates([rate_pd[0].parameters])
        self.assertEqual(len(rate_i), 16)
        self.assertEqual(rate_i[0].parameters["atc_rat_pit_i"], 0.0)
        self.assertEqual(rate_i[-1].parameters["atc_rat_pit_i"], 0.03)

        attitude = attitude_p_candidates(rate_i[0].parameters)
        self.assertEqual(len(attitude), 17)
        self.assertEqual(attitude[0].parameters["atc_ang_pit_p"], 2.0)
        self.assertEqual(attitude[-1].parameters["atc_ang_pit_p"], 10.0)

        loiter = loiter_xy_candidates(attitude[0].parameters)
        self.assertEqual(len(loiter), 15 * 24)
        self.assertEqual(loiter[0].parameters["psc_ne_pos_p"], 0.1)
        self.assertEqual(loiter[0].parameters["psc_ne_vel_p"], 0.2)
        self.assertEqual(loiter[-1].parameters["psc_ne_pos_p"], 1.5)
        self.assertEqual(loiter[-1].parameters["psc_ne_vel_p"], 2.5)

        gains = moving_mass_gain_candidates(loiter[0].parameters)
        self.assertEqual(len(gains), 33)
        self.assertEqual(gains[0].parameters["moving_mass_assist_gain_m_per_Nm"], 0.0)
        self.assertEqual(gains[-1].parameters["moving_mass_assist_gain_m_per_Nm"], 0.08)

    def test_candidate_order_and_keys_are_deterministic_and_unique(self):
        first = rate_pd_coarse_candidates()
        second = rate_pd_coarse_candidates()
        self.assertEqual([candidate.key for candidate in first], [candidate.key for candidate in second])
        self.assertEqual(len({candidate.key for candidate in first}), len(first))
        self.assertEqual(len({candidate.candidate_id for candidate in first}), len(first))

    def test_loiter_grid_excludes_inactive_velocity_i_and_d(self):
        candidate = loiter_xy_candidates(
            {
                "atc_rat_pit_p": 0.035,
                "atc_rat_pit_i": 0.01,
                "atc_rat_pit_d": 0.002,
                "atc_ang_pit_p": 7.0,
            },
            quick=True,
        )[0]
        self.assertNotIn("psc_ne_vel_i", candidate.parameters)
        self.assertNotIn("psc_ne_vel_d", candidate.parameters)

    def test_tail_metric_helpers(self):
        self.assertEqual(zero_crossing_count([1.0, 0.0, -1.0, -0.5, 1.0], 1e-9), 2)
        self.assertEqual(direction_reversal_count([0.0, 1.0, 2.0, 1.0, 0.0, 1.0], 1e-9), 2)
        self.assertAlmostEqual(path_length([0.0, 3.0, 3.0], [0.0, 0.0, 4.0]), 7.0)

    def test_normalized_score_uses_documented_reference_scales(self):
        row = {spec.metric: spec.reference_scale for spec in SCORE_SPECS["rate_pd"]}
        score, components = normalized_score(row, "rate_pd")
        self.assertAlmostEqual(score, 1.0)
        self.assertTrue(all(value == 1.0 for value in json.loads(components).values()))

    def test_top_n_excludes_hard_rejections(self):
        rows = [
            {"candidate_key": "a", "normalized_score": 0.1, "rejected": True},
            {"candidate_key": "b", "normalized_score": 0.2, "rejected": False},
            {"candidate_key": "c", "normalized_score": 0.3, "rejected": False},
        ]
        self.assertEqual([row["candidate_key"] for row in top_candidates(rows, 1)], ["b"])

    def test_resume_store_rejects_duplicate_run_keys(self):
        with tempfile.TemporaryDirectory() as tmp:
            store = ScenarioResultStore(
                Path(tmp) / "rows.csv", cache_fingerprint="fingerprint"
            )
            row = {"run_key": "same", "cache_fingerprint": "fingerprint"}
            store.add(row)
            with self.assertRaisesRegex(ValueError, "duplicate run key"):
                store.add(row)

    def test_hard_rejection_flags_ground_contact(self):
        rows = [
            {
                "time": index * 0.1,
                "min_body_z": -0.01,
                "theta": 0.0,
                "omega": 0.0,
                "rate_error": 0.0,
                "x_error": 0.0,
                "x": 0.0,
                "z": 0.0,
                "vx": 0.0,
            }
            for index in range(3)
        ]
        scenario = LoiterScenarioConfig(name="ground_contact", duration_s=0.3)
        result = LoiterRunResult("<synthetic>", scenario, rows, {}, False, "")
        metrics = compute_scenario_metrics("loiter_xy", result, tail_window_s=0.2)
        self.assertTrue(metrics["rejected"])
        self.assertIn("ground contact", metrics["rejection_reasons"])


class ControllerGridCacheAndValidityTests(unittest.TestCase):
    def _fingerprint_options(self, root: Path, *, quick: bool = False, tail: float = 2.0) -> WorkflowOptions:
        vane = root / "vane.json"
        moving = root / "moving.json"
        if not vane.exists():
            vane.write_text('{"source":"vane"}\n', encoding="utf-8")
        if not moving.exists():
            moving.write_text('{"source":"moving"}\n', encoding="utf-8")
        return WorkflowOptions(
            output_dir=root / "output",
            profile_output_dir=root / "profiles",
            vane_param_source=vane,
            moving_mass_param_source=moving,
            quick=quick,
            tail_window_s=tail,
        )

    def test_workflow_fingerprint_tracks_schema_sources_mode_and_tail_policy(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            options = self._fingerprint_options(root)
            first = build_workflow_fingerprint(options)
            self.assertEqual(first.digest, build_workflow_fingerprint(options).digest)
            self.assertNotEqual(
                first.digest,
                build_workflow_fingerprint(
                    options, cache_schema_version=CACHE_SCHEMA_VERSION + 1
                ).digest,
            )
            self.assertNotEqual(
                first.digest,
                build_workflow_fingerprint(replace(options, quick=True)).digest,
            )
            self.assertNotEqual(
                first.digest,
                build_workflow_fingerprint(replace(options, tail_window_s=1.5)).digest,
            )
            Path(options.vane_param_source).write_text(
                '{"source":"changed"}\n', encoding="utf-8"
            )
            self.assertNotEqual(first.digest, build_workflow_fingerprint(options).digest)

    def test_moving_mass_scenario_dependency_changes_workflow_fingerprint(self):
        self.assertIn("analysis/moving_mass_comparison.py", WORKFLOW_IMPLEMENTATION_FILES)
        with tempfile.TemporaryDirectory() as tmp:
            options = self._fingerprint_options(Path(tmp))
            original = build_workflow_fingerprint(options)
            real_sha256_file = controller_grid_search_module._sha256_file

            def changed_dependency_hash(path: Path) -> str:
                if path.as_posix().endswith("analysis/moving_mass_comparison.py"):
                    return "changed-moving-mass-scenario-dependency"
                return real_sha256_file(path)

            with patch.object(
                controller_grid_search_module,
                "_sha256_file",
                side_effect=changed_dependency_hash,
            ):
                changed = build_workflow_fingerprint(options)
            self.assertNotEqual(original.digest, changed.digest)

    def test_profile_output_defaults_and_explicit_destinations(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            full = WorkflowOptions(output_dir=root / "full")
            quick = replace(full, output_dir=root / "quick", quick=True)
            explicit = root / "explicit-profiles"
            self.assertEqual(resolve_profile_output_directory(full), Path("params"))
            self.assertEqual(
                resolve_profile_output_directory(quick),
                root / "quick" / "profiles",
            )
            self.assertEqual(
                resolve_profile_output_directory(replace(full, profile_output_dir=explicit)),
                explicit,
            )
            self.assertEqual(
                resolve_profile_output_directory(replace(quick, profile_output_dir=explicit)),
                explicit,
            )

    def test_selected_parameter_boundaries_are_machine_readable(self):
        stage_aggregates = {
            "rate_pd": [
                {"rejected": False, "atc_rat_pit_p": 0.07, "atc_rat_pit_d": 0.008},
                {"rejected": False, "atc_rat_pit_p": 0.005, "atc_rat_pit_d": 0.0},
                {"rejected": False, "atc_rat_pit_p": 0.08, "atc_rat_pit_d": 0.004},
            ],
            "rate_i": [
                {"rejected": False, "atc_rat_pit_p": 0.07, "atc_rat_pit_i": 0.0, "atc_rat_pit_d": 0.008},
                {"rejected": False, "atc_rat_pit_i": 0.03},
            ],
            "attitude_p": [
                {"rejected": False, "atc_ang_pit_p": 10.0},
                {"rejected": False, "atc_ang_pit_p": 2.0},
            ],
            "loiter_xy": [
                {"rejected": False, "psc_ne_pos_p": 0.5, "psc_ne_vel_p": 0.9},
                {"rejected": False, "psc_ne_pos_p": 0.1, "psc_ne_vel_p": 0.2},
                {"rejected": False, "psc_ne_pos_p": 1.5, "psc_ne_vel_p": 2.5},
            ],
            "moving_mass_gain": [
                {"rejected": False, "moving_mass_assist_gain_m_per_Nm": 0.055},
                {"rejected": False, "moving_mass_assist_gain_m_per_Nm": 0.0},
                {"rejected": False, "moving_mass_assist_gain_m_per_Nm": 0.08},
            ],
        }
        diagnostics = {
            row["parameter"]: row
            for row in selected_parameter_boundary_rows(stage_aggregates)
        }
        self.assertTrue(diagnostics["atc_rat_pit_d"]["selected_at_upper_boundary"])
        self.assertTrue(diagnostics["atc_ang_pit_p"]["selected_at_upper_boundary"])
        self.assertTrue(diagnostics["atc_rat_pit_i"]["selected_at_lower_boundary"])
        self.assertEqual(diagnostics["atc_rat_pit_d"]["parameter_search_max"], 0.008)
        self.assertIn("not evidence of an interior or global optimum", diagnostics["atc_rat_pit_d"]["boundary_warning"])

    def test_identical_fingerprint_reuses_rows_and_changed_schema_invalidates(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            options = self._fingerprint_options(root)
            current = build_workflow_fingerprint(options)
            changed = build_workflow_fingerprint(
                options, cache_schema_version=CACHE_SCHEMA_VERSION + 1
            )
            path = root / "scenario_results.csv"
            store = ScenarioResultStore(path, cache_fingerprint=current.digest)
            row = {"run_key": "same", "cache_fingerprint": current.digest}
            store.add(row)
            resumed = ScenarioResultStore(path, cache_fingerprint=current.digest)
            self.assertEqual(resumed.get("same")["run_key"], "same")
            with self.assertRaisesRegex(ValueError, "stale scenario cache fingerprint"):
                ScenarioResultStore(path, cache_fingerprint=changed.digest)

    def test_stale_previous_stage_metadata_is_rejected(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            options = self._fingerprint_options(root)
            output = Path(options.output_dir)
            output.mkdir(parents=True)
            (output / "01_rate_pd_all.csv").write_text("stage\n", encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "stale workflow fingerprint"):
                validate_previous_stage_metadata(
                    output,
                    {"workflow_fingerprint": "outdated"},
                    build_workflow_fingerprint(options),
                )

    def test_stage_specific_tail_windows(self):
        rate_scenarios = rate_pd_scenarios(False)
        self.assertEqual(rate_scenarios[0].tail_window_s, 0.75)
        self.assertEqual(rate_scenarios[0].config.duration_s, 3.0)
        self.assertTrue(rate_scenarios[0].primary_score)
        self.assertEqual(rate_scenarios[2].config.duration_s, 1.5)
        self.assertTrue(rate_scenarios[2].primary_score)
        self.assertFalse(rate_scenarios[4].primary_score)
        self.assertEqual(rate_i_scenarios(False)[2].tail_window_s, 1.0)
        attitude = attitude_p_scenarios(False)[0]
        self.assertEqual(attitude.tail_window_s, 1.0)
        self.assertEqual(attitude.config.duration_s, 5.0)
        self.assertIsNone(loiter_xy_scenarios(False)[0].tail_window_s)

    def test_oversized_tail_window_cannot_cover_the_complete_run(self):
        rows = [
            {
                "time": float(index),
                "rate_error": math.radians(value),
                "min_body_z": 1.0,
            }
            for index, value in enumerate((12.0, 0.0, 0.0))
        ]
        scenario = LoiterScenarioConfig(name="short_rate", mode="RATE", duration_s=2.0)
        result = LoiterRunResult("<synthetic>", scenario, rows, {"pass": True}, False, "")
        metrics = compute_scenario_metrics("rate_pd", result, tail_window_s=10.0)
        self.assertGreater(metrics["rms_rate_error_deg_s"], 0.0)
        self.assertEqual(metrics["tail_rms_rate_error_deg_s"], 0.0)

    def test_required_rate_and_attitude_validity_gates(self):
        rate = SearchScenario(
            LoiterScenarioConfig(name="rate_required"), validity_gate="settled"
        )
        self.assertIn(
            "did not settle",
            stage_validity_reasons("rate_pd", rate, {"settled": False})[0],
        )
        attitude = SearchScenario(
            LoiterScenarioConfig(name="attitude_required"),
            validity_gate="attitude_terminal",
        )
        reasons = stage_validity_reasons(
            "attitude_p",
            attitude,
            {
                "settled": False,
                "terminal_abs_theta_deg": 8.0,
                "terminal_abs_omega_deg_s": 13.0,
            },
        )
        self.assertIn("neither settled", reasons[0])

        bias = SearchScenario(
            LoiterScenarioConfig(name="rate_bias"), validity_gate="rate_bias"
        )
        self.assertEqual(
            stage_validity_reasons(
                "rate_i",
                bias,
                {
                    "tail_mean_abs_rate_error_deg_s": 40.0,
                    "terminal_abs_rate_error_deg_s": 30.0,
                },
            ),
            [],
        )
        self.assertTrue(
            stage_validity_reasons(
                "rate_i",
                bias,
                {
                    "tail_mean_abs_rate_error_deg_s": 40.1,
                    "terminal_abs_rate_error_deg_s": 30.1,
                },
            )
        )

    def test_robustness_only_unsettled_case_does_not_invalidate(self):
        robustness = SearchScenario(
            LoiterScenarioConfig(name="robustness"),
            primary_score=False,
            validity_gate="",
        )
        self.assertEqual(
            stage_validity_reasons("rate_pd", robustness, {"settled": False}), []
        )

    def test_no_valid_candidate_blocks_selection_and_profile_publication(self):
        rejected = [{"candidate_key": "bad", "rejected": True}]
        with self.assertRaisesRegex(RuntimeError, "no valid candidate"):
            require_best_aggregate_row("rate_pd", rejected)
        with tempfile.TemporaryDirectory() as tmp:
            with self.assertRaisesRegex(RuntimeError, "no valid candidate"):
                write_profiles(
                    {stage: rejected for stage in ("rate_pd", "rate_i", "attitude_p", "loiter_xy", "moving_mass_gain")},
                    vane_source=Path("params/loiter_example.json"),
                    moving_mass_source=Path("params/moving_mass_prototype_2kg.json"),
                    output_directory=Path(tmp),
                )


class ControllerGridQuickEndToEndTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls._tmp = tempfile.TemporaryDirectory()
        root = Path(cls._tmp.name)
        cls.output_dir = root / "output"
        cls.profile_dir = cls.output_dir / "profiles"
        cls.canonical_profile_paths = (
            Path("params/loiter_tuned_vane_only.json"),
            Path("params/moving_mass_prototype_2kg_tuned.json"),
        )
        cls.canonical_hashes_before = {
            path: hashlib.sha256(path.read_bytes()).hexdigest()
            for path in cls.canonical_profile_paths
        }
        cls.options = WorkflowOptions(
                stage="all",
                output_dir=cls.output_dir,
                quick=True,
                resume=False,
            )
        cls.result = run_workflow(cls.options)
        cls.resumed_result = run_workflow(replace(cls.options, resume=True))

    @classmethod
    def tearDownClass(cls):
        cls._tmp.cleanup()

    def test_quick_end_to_end_outputs_and_unique_keys(self):
        self.assertTrue(self.result["workbook_path"].exists())
        self.assertTrue(self.result["markdown_path"].exists())
        self.assertEqual(
            len({row["run_key"] for row in self.result["scenario_rows"]}),
            len(self.result["scenario_rows"]),
        )
        for stage in ("rate_pd", "rate_i", "attitude_p", "loiter_xy", "moving_mass_gain"):
            self.assertGreater(len(self.result["stage_aggregates"][stage]), 0)
        self.assertEqual(
            self.resumed_result["metadata"]["resume_skipped_scenario_count"],
            len(self.result["scenario_rows"]),
        )
        self.assertEqual(
            self.result["metadata"]["workflow_fingerprint"],
            self.resumed_result["metadata"]["workflow_fingerprint"],
        )

    def test_workbook_sheet_names_and_row_counts(self):
        workbook = load_workbook(self.result["workbook_path"], read_only=False, data_only=False)
        self.assertEqual(tuple(workbook.sheetnames), SHEET_NAMES)
        aggregates = self.result["stage_aggregates"]
        expected = {
            "01_rate_pd_all": len(aggregates["rate_pd"]),
            "02_rate_pd_top50": len(top_candidates(aggregates["rate_pd"], 50)),
            "03_rate_i_all": len(aggregates["rate_i"]),
            "04_attitude_p_all": len(aggregates["attitude_p"]),
            "05_loiter_xy_all": len(aggregates["loiter_xy"]),
            "06_loiter_xy_top50": len(top_candidates(aggregates["loiter_xy"], 50)),
            "07_moving_mass_gain_all": len(aggregates["moving_mass_gain"]),
            "08_scenario_summary": len(self.result["scenario_rows"]),
            "09_best_parameters": len(best_parameter_rows(aggregates)),
            "10_metadata": len(self.result["metadata"]),
        }
        for sheet_name, row_count in expected.items():
            self.assertEqual(workbook[sheet_name].max_row, row_count + 1, sheet_name)
            self.assertEqual(workbook[sheet_name].freeze_panes, "A2")
            self.assertIsNotNone(workbook[sheet_name].auto_filter.ref)

    def test_profiles_preserve_structure_and_are_loadable(self):
        vane_path = self.result["vane_profile_path"]
        moving_path = self.result["moving_mass_profile_path"]
        vane = json.loads(vane_path.read_text(encoding="utf-8"))
        moving = json.loads(moving_path.read_text(encoding="utf-8"))
        self.assertEqual(set(vane), {"rigid_body", "interactive", "controller"})
        self.assertEqual(set(moving), {"rigid_body", "interactive", "controller", "analysis"})
        self.assertIn("moving_mass_assist_gain_m_per_Nm", moving["analysis"])
        self.assertNotIn("psc_ne_vel_i", vane["controller"])
        self.assertNotIn("psc_ne_vel_d", vane["controller"])
        load_interactive_config(vane_path)
        load_interactive_config(moving_path)

    def test_quick_default_uses_output_profiles_and_preserves_canonical_hashes(self):
        self.assertEqual(
            self.result["vane_profile_path"],
            self.profile_dir / "loiter_tuned_vane_only.json",
        )
        self.assertEqual(
            self.result["moving_mass_profile_path"],
            self.profile_dir / "moving_mass_prototype_2kg_tuned.json",
        )
        hashes_after = {
            path: hashlib.sha256(path.read_bytes()).hexdigest()
            for path in self.canonical_profile_paths
        }
        self.assertEqual(self.canonical_hashes_before, hashes_after)


if __name__ == "__main__":
    unittest.main()
