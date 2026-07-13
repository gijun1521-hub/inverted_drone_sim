from __future__ import annotations

import csv
import hashlib
import json
import math
import subprocess
import time
from dataclasses import asdict, dataclass, replace
from decimal import Decimal
from pathlib import Path
from typing import Any, Sequence

import numpy as np

try:
    from .headless_loiter import (
        LoiterRunResult,
        LoiterScenarioConfig,
        authority_stress_scenario,
        run_headless_loiter,
        scenario_by_name,
    )
    from .moving_mass_comparison import DEFAULT_SCENARIOS
    from ..params import load_interactive_config
except ImportError:  # pragma: no cover - supports top-level script execution
    from analysis.headless_loiter import (
        LoiterRunResult,
        LoiterScenarioConfig,
        authority_stress_scenario,
        run_headless_loiter,
        scenario_by_name,
    )
    from analysis.moving_mass_comparison import DEFAULT_SCENARIOS
    from params import load_interactive_config


STAGES = ("rate_pd", "rate_i", "attitude_p", "loiter_xy", "moving_mass_gain")
SHEET_NAMES = (
    "01_rate_pd_all",
    "02_rate_pd_top50",
    "03_rate_i_all",
    "04_attitude_p_all",
    "05_loiter_xy_all",
    "06_loiter_xy_top50",
    "07_moving_mass_gain_all",
    "08_scenario_summary",
    "09_best_parameters",
    "10_metadata",
)

PARAMETER_FIELDS = (
    "atc_rat_pit_p",
    "atc_rat_pit_i",
    "atc_rat_pit_d",
    "atc_ang_pit_p",
    "psc_ne_pos_p",
    "psc_ne_vel_p",
    "moving_mass_assist_gain_m_per_Nm",
)
CONTROLLER_PARAMETER_FIELDS = PARAMETER_FIELDS[:-1]
EFFECTIVE_FIELDS = {name: f"effective_{name}" for name in CONTROLLER_PARAMETER_FIELDS}

SCENARIO_METRIC_FIELDS = (
    "duration_s",
    "analytical_pass",
    "rms_rate_error_deg_s",
    "tail_rms_rate_error_deg_s",
    "tail_mean_abs_rate_error_deg_s",
    "rate_overshoot_deg_s",
    "rate_zero_crossings",
    "rms_theta_deg",
    "tail_rms_theta_deg",
    "theta_overshoot_deg",
    "tail_peak_to_peak_theta_deg",
    "max_omega_deg_s",
    "tail_rms_x_m",
    "tail_rms_vx_m_s",
    "tail_peak_to_peak_x_m",
    "tail_path_length_m",
    "x_error_zero_crossings",
    "oscillation_peak_count",
    "settling_time_s",
    "settled",
    "final_abs_x_error_m",
    "rms_x_error_m",
    "vane_command_rms_deg",
    "vane_command_max_deg",
    "vane_angle_saturation_percent",
    "servo_rate_saturation_percent",
    "mixer_saturation_percent",
    "control_effort_rms_Nm",
    "integrator_rms_Nm",
    "integrator_max_abs_Nm",
    "integrator_inhibition_percent",
    "anti_windup_rms_Nm",
    "moving_mass_max_offset_m",
    "moving_mass_saturation_percent",
    "moving_mass_tracking_rms_m",
    "moving_mass_target_reversals",
    "moving_mass_actual_reversals",
    "baseline_rms_theta_deg",
    "baseline_rms_x_error_m",
    "baseline_final_abs_x_error_m",
    "baseline_tail_path_length_m",
    "baseline_vane_command_rms_deg",
    "rms_theta_improvement_deg",
    "rms_theta_ratio",
    "rms_x_change_m",
    "rms_x_ratio",
    "final_x_ratio",
    "tail_path_ratio",
    "vane_command_rms_ratio",
    "ground_contact",
    "attitude_limit_exceeded",
    "unbounded_growth",
    "excessive_sustained_saturation",
    "effective_parameter_mismatch",
)

SCENARIO_CSV_FIELDS = (
    "run_key",
    "stage",
    "candidate_id",
    "candidate_key",
    "search_phase",
    "scenario_name",
    "symmetry_group",
    "primary_score",
    "param_file",
    *PARAMETER_FIELDS,
    "pass",
    "rejected",
    "rejection_reasons",
    "crash_reason",
    *SCENARIO_METRIC_FIELDS,
    *EFFECTIVE_FIELDS.values(),
)

AGGREGATE_METRIC_FIELDS = tuple(
    name
    for name in SCENARIO_METRIC_FIELDS
    if name
    not in {
        "settled",
        "analytical_pass",
        "ground_contact",
        "attitude_limit_exceeded",
        "unbounded_growth",
        "excessive_sustained_saturation",
        "effective_parameter_mismatch",
    }
) + ("left_right_symmetry_error",)

AGGREGATE_CSV_FIELDS = (
    "stage",
    "candidate_id",
    "candidate_key",
    "search_phase",
    *PARAMETER_FIELDS,
    "scenario_count",
    "primary_scenario_count",
    "missing_scenario_count",
    "rejected",
    "rejection_reasons",
    "normalized_score",
    "rank",
    "selection_label",
    *AGGREGATE_METRIC_FIELDS,
)


@dataclass(frozen=True)
class Candidate:
    stage: str
    parameters: dict[str, float]
    search_phase: str = "full"

    @property
    def key(self) -> str:
        payload = ",".join(
            f"{name}={_format_number(self.parameters[name])}"
            for name in sorted(self.parameters)
        )
        return f"{self.stage}|{payload}"

    @property
    def candidate_id(self) -> str:
        digest = hashlib.sha1(self.key.encode("utf-8")).hexdigest()[:10]
        return f"{self.stage}-{digest}"


@dataclass(frozen=True)
class SearchScenario:
    config: LoiterScenarioConfig
    primary_score: bool = True
    symmetry_group: str = ""


@dataclass(frozen=True)
class ScoreSpec:
    metric: str
    reference_scale: float
    weight: float


SCORE_SPECS: dict[str, tuple[ScoreSpec, ...]] = {
    "rate_pd": (
        ScoreSpec("rms_rate_error_deg_s", 35.0, 0.19),
        ScoreSpec("tail_rms_rate_error_deg_s", 4.0, 0.25),
        ScoreSpec("rate_overshoot_deg_s", 10.0, 0.08),
        ScoreSpec("settling_time_s", 2.0, 0.12),
        ScoreSpec("rate_zero_crossings", 5.0, 0.05),
        ScoreSpec("vane_command_rms_deg", 6.0, 0.08),
        ScoreSpec("vane_angle_saturation_percent", 10.0, 0.05),
        ScoreSpec("servo_rate_saturation_percent", 10.0, 0.08),
        ScoreSpec("control_effort_rms_Nm", 0.12, 0.10),
    ),
    "rate_i": (
        ScoreSpec("tail_mean_abs_rate_error_deg_s", 1.0, 0.28),
        ScoreSpec("tail_rms_rate_error_deg_s", 2.0, 0.12),
        ScoreSpec("rate_overshoot_deg_s", 8.0, 0.09),
        ScoreSpec("settling_time_s", 2.5, 0.10),
        ScoreSpec("integrator_rms_Nm", 0.05, 0.09),
        ScoreSpec("integrator_inhibition_percent", 10.0, 0.08),
        ScoreSpec("anti_windup_rms_Nm", 0.03, 0.07),
        ScoreSpec("vane_angle_saturation_percent", 10.0, 0.07),
        ScoreSpec("control_effort_rms_Nm", 0.12, 0.10),
    ),
    "attitude_p": (
        ScoreSpec("rms_theta_deg", 8.0, 0.16),
        ScoreSpec("tail_rms_theta_deg", 1.0, 0.24),
        ScoreSpec("theta_overshoot_deg", 3.0, 0.09),
        ScoreSpec("settling_time_s", 2.5, 0.12),
        ScoreSpec("tail_peak_to_peak_theta_deg", 2.0, 0.12),
        ScoreSpec("max_omega_deg_s", 100.0, 0.07),
        ScoreSpec("vane_command_rms_deg", 6.0, 0.07),
        ScoreSpec("vane_angle_saturation_percent", 10.0, 0.05),
        ScoreSpec("left_right_symmetry_error", 0.10, 0.08),
    ),
    "loiter_xy": (
        ScoreSpec("tail_rms_x_m", 0.10, 0.17),
        ScoreSpec("tail_rms_theta_deg", 1.0, 0.10),
        ScoreSpec("tail_rms_vx_m_s", 0.10, 0.14),
        ScoreSpec("tail_peak_to_peak_x_m", 0.20, 0.14),
        ScoreSpec("tail_path_length_m", 0.25, 0.18),
        ScoreSpec("settling_time_s", 6.0, 0.07),
        ScoreSpec("final_abs_x_error_m", 0.15, 0.06),
        ScoreSpec("rms_x_error_m", 0.70, 0.05),
        ScoreSpec("vane_command_rms_deg", 6.0, 0.04),
        ScoreSpec("servo_rate_saturation_percent", 10.0, 0.03),
        ScoreSpec("left_right_symmetry_error", 0.10, 0.02),
    ),
    "moving_mass_gain": (
        ScoreSpec("rms_theta_ratio", 1.0, 0.30),
        ScoreSpec("rms_x_ratio", 1.0, 0.17),
        ScoreSpec("final_x_ratio", 1.0, 0.10),
        ScoreSpec("tail_path_ratio", 1.0, 0.13),
        ScoreSpec("vane_command_rms_ratio", 1.0, 0.10),
        ScoreSpec("moving_mass_tracking_rms_m", 0.005, 0.08),
        ScoreSpec("moving_mass_saturation_percent", 10.0, 0.07),
        ScoreSpec("moving_mass_max_offset_m", 0.05, 0.05),
    ),
}


def _format_number(value: float) -> str:
    return f"{float(value):.12g}"


def decimal_grid(start: str, stop: str, step: str) -> list[float]:
    start_d, stop_d, step_d = Decimal(start), Decimal(stop), Decimal(step)
    count = int((stop_d - start_d) / step_d)
    values = [float(start_d + step_d * index) for index in range(count + 1)]
    if Decimal(str(values[-1])) != stop_d:
        raise ValueError(f"grid does not end exactly at {stop}")
    return values


def rate_pd_coarse_candidates(quick: bool = False) -> list[Candidate]:
    p_values = [0.020, 0.035, 0.050] if quick else decimal_grid("0.005", "0.080", "0.005")
    d_values = [0.001, 0.002, 0.003] if quick else decimal_grid("0.0000", "0.0080", "0.0005")
    return [
        Candidate(
            "rate_pd",
            {"atc_rat_pit_p": p, "atc_rat_pit_i": 0.0, "atc_rat_pit_d": d},
            "coarse",
        )
        for p in p_values
        for d in d_values
    ]


def rate_pd_fine_candidates(
    top_coarse: Sequence[dict[str, Any]],
    *,
    existing_keys: set[str] | None = None,
) -> list[Candidate]:
    existing = set(existing_keys or ())
    candidates: dict[str, Candidate] = {}
    for row in top_coarse:
        p_center = _as_float(row.get("atc_rat_pit_p"))
        d_center = _as_float(row.get("atc_rat_pit_d"))
        for p_delta in (-0.0025, 0.0, 0.0025):
            for d_delta in (-0.00025, 0.0, 0.00025):
                p = round(p_center + p_delta, 7)
                d = round(d_center + d_delta, 8)
                if not (0.005 <= p <= 0.080 and 0.0 <= d <= 0.008):
                    continue
                candidate = Candidate(
                    "rate_pd",
                    {"atc_rat_pit_p": p, "atc_rat_pit_i": 0.0, "atc_rat_pit_d": d},
                    "fine",
                )
                if candidate.key not in existing:
                    candidates[candidate.key] = candidate
    return [candidates[key] for key in sorted(candidates)]


def rate_i_candidates(
    pd_rows: Sequence[dict[str, Any]], quick: bool = False
) -> list[Candidate]:
    i_values = [0.0, 0.010, 0.020] if quick else decimal_grid("0.000", "0.030", "0.002")
    candidates: dict[str, Candidate] = {}
    for pd_row in pd_rows:
        for i_value in i_values:
            candidate = Candidate(
                "rate_i",
                {
                    "atc_rat_pit_p": _as_float(pd_row["atc_rat_pit_p"]),
                    "atc_rat_pit_i": i_value,
                    "atc_rat_pit_d": _as_float(pd_row["atc_rat_pit_d"]),
                },
            )
            candidates[candidate.key] = candidate
    return [candidates[key] for key in sorted(candidates)]


def attitude_p_candidates(rate_parameters: dict[str, float], quick: bool = False) -> list[Candidate]:
    values = [3.5, 7.0, 9.0] if quick else decimal_grid("2.0", "10.0", "0.5")
    return [
        Candidate("attitude_p", {**rate_parameters, "atc_ang_pit_p": value})
        for value in values
    ]


def loiter_xy_candidates(attitude_parameters: dict[str, float], quick: bool = False) -> list[Candidate]:
    pos_values = [0.4, 0.8, 1.2] if quick else decimal_grid("0.1", "1.5", "0.1")
    vel_values = [0.6, 1.1, 1.8] if quick else decimal_grid("0.2", "2.5", "0.1")
    return [
        Candidate(
            "loiter_xy",
            {**attitude_parameters, "psc_ne_pos_p": pos_p, "psc_ne_vel_p": vel_p},
        )
        for pos_p in pos_values
        for vel_p in vel_values
    ]


def moving_mass_gain_candidates(controller_parameters: dict[str, float], quick: bool = False) -> list[Candidate]:
    values = [0.0, 0.025, 0.050] if quick else decimal_grid("0.000", "0.080", "0.0025")
    return [
        Candidate(
            "moving_mass_gain",
            {**controller_parameters, "moving_mass_assist_gain_m_per_Nm": value},
        )
        for value in values
    ]


def _duration(value: float, quick: bool) -> float:
    return 0.30 if quick else value


def rate_pd_scenarios(quick: bool = False) -> list[SearchScenario]:
    duration = _duration(1.5, quick)
    scenarios = []
    for label, omega in (("pos_moderate", 60.0), ("neg_moderate", -60.0), ("pos_strong", 120.0), ("neg_strong", -120.0)):
        scenarios.append(
            SearchScenario(
                LoiterScenarioConfig(
                    name=f"rate_{label}", mode="RATE", duration_s=duration,
                    initial_z=6.0,
                    initial_omega_deg_s=omega, capture_current_target=True,
                    max_theta_deg_limit=90.0, max_saturation_percent=100.0,
                ),
                symmetry_group=label.replace("pos_", "").replace("neg_", ""),
            )
        )
    for label, omega in (("pos", 60.0), ("neg", -60.0)):
        scenarios.append(
            SearchScenario(
                LoiterScenarioConfig(
                    name=f"rate_low_authority_{label}", mode="RATE", duration_s=duration,
                    initial_z=6.0,
                    initial_omega_deg_s=omega, capture_current_target=True,
                    vane_angle_max_deg=8.0, vane_rate_limit_deg_s=90.0,
                    max_theta_deg_limit=100.0, max_saturation_percent=100.0,
                ),
                symmetry_group="low_authority",
            )
        )
    return scenarios


def rate_i_scenarios(quick: bool = False) -> list[SearchScenario]:
    recovery_duration = _duration(1.5, quick)
    bias_duration = _duration(4.0, quick)
    scenarios = [
        SearchScenario(
            LoiterScenarioConfig(
                name=f"rate_i_{label}", mode="RATE", duration_s=recovery_duration,
                initial_z=6.0,
                initial_omega_deg_s=omega, capture_current_target=True,
                max_theta_deg_limit=90.0, max_saturation_percent=100.0,
            ),
            symmetry_group="moderate",
        )
        for label, omega in (("pos_moderate", 60.0), ("neg_moderate", -60.0))
    ]
    for label, moment in (("positive_bias", 0.005), ("negative_bias", -0.005)):
        scenarios.append(
            SearchScenario(
                LoiterScenarioConfig(
                    name=f"rate_i_{label}", mode="LOITER", duration_s=bias_duration,
                    initial_z=1.0,
                    disturbance_start_s=0.0, disturbance_duration_s=bias_duration,
                    disturbance_moment=moment, capture_current_target=True,
                    max_theta_deg_limit=90.0, max_saturation_percent=100.0,
                ),
                symmetry_group="persistent_bias",
            )
        )
    return scenarios


def attitude_p_scenarios(quick: bool = False) -> list[SearchScenario]:
    duration = _duration(4.0, quick)
    scenarios = []
    for angle in (5, -5, 10, -10, 15, -15):
        scenarios.append(
            SearchScenario(
                LoiterScenarioConfig(
                    name=f"attitude_{angle:+d}deg", mode="STABILIZE", duration_s=duration,
                    initial_z=6.0,
                    initial_theta_deg=float(angle), capture_current_target=True,
                    max_theta_deg_limit=80.0, max_saturation_percent=100.0,
                ),
                symmetry_group=f"attitude_{abs(angle)}deg",
            )
        )
    return scenarios


def loiter_xy_scenarios(quick: bool = False) -> list[SearchScenario]:
    duration = _duration(8.0, quick)
    scenarios = [
        SearchScenario(
            LoiterScenarioConfig(name="upright_hold", duration_s=duration, capture_current_target=False),
            symmetry_group="upright",
        )
    ]
    for label, initial_x in (("right", 1.2), ("left", -1.2)):
        scenarios.append(
            SearchScenario(
                replace(
                    scenario_by_name("initial_x_offset_recovery", duration),
                    name=f"initial_x_offset_{label}", initial_x=initial_x,
                ),
                symmetry_group="initial_x_offset",
            )
        )
    for label, force in (("right", 8.0), ("left", -8.0)):
        scenarios.append(
            SearchScenario(
                replace(
                    scenario_by_name("horizontal_impulse_recovery", duration),
                    name=f"horizontal_impulse_{label}", disturbance_force_x=force,
                ),
                symmetry_group="horizontal_impulse",
            )
        )
    for label, stick in (("right", 0.65), ("left", -0.65)):
        scenarios.append(
            SearchScenario(
                replace(
                    scenario_by_name("stick_move_release", duration),
                    name=f"stick_move_release_{label}", stick_x=stick,
                ),
                symmetry_group="stick_move_release",
            )
        )
    stress = authority_stress_scenario(duration)
    for label, sign in (("right", 1.0), ("left", -1.0)):
        scenarios.append(
            SearchScenario(
                replace(
                    stress,
                    name=f"authority_stress_{label}",
                    initial_x=sign * abs(stress.initial_x),
                    disturbance_force_x=sign * abs(stress.disturbance_force_x),
                ),
                primary_score=False,
                symmetry_group="authority_stress",
            )
        )
    return scenarios


def moving_mass_gain_scenarios(quick: bool = False) -> list[SearchScenario]:
    duration = 0.30 if quick else None
    return [SearchScenario(scenario_by_name(name, duration)) for name in DEFAULT_SCENARIOS]


def _as_float(value: Any, default: float = 0.0) -> float:
    if value in (None, ""):
        return default
    return float(value)


def _as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes"}


def _rms(values: np.ndarray) -> float:
    return float(np.sqrt(np.mean(values * values))) if values.size else 0.0


def zero_crossing_count(values: Sequence[float], tolerance: float = 0.0) -> int:
    array = np.asarray(values, dtype=float)
    signs = np.sign(array[np.abs(array) > tolerance])
    return int(np.count_nonzero(signs[1:] != signs[:-1])) if signs.size > 1 else 0


def direction_reversal_count(values: Sequence[float], tolerance: float = 0.0) -> int:
    array = np.asarray(values, dtype=float)
    if array.size < 3:
        return 0
    return zero_crossing_count(np.diff(array), tolerance)


def path_length(x: Sequence[float], z: Sequence[float]) -> float:
    x_array, z_array = np.asarray(x, dtype=float), np.asarray(z, dtype=float)
    if x_array.size < 2:
        return 0.0
    return float(np.sum(np.hypot(np.diff(x_array), np.diff(z_array))))


def oscillation_peak_count(values: Sequence[float], threshold: float) -> int:
    array = np.abs(np.asarray(values, dtype=float))
    if array.size < 3:
        return 0
    peaks = (array[1:-1] > array[:-2]) & (array[1:-1] >= array[2:]) & (array[1:-1] > threshold)
    return int(np.count_nonzero(peaks))


def _settling_time(times: np.ndarray, condition: np.ndarray) -> tuple[float, bool]:
    if not times.size:
        return 0.0, False
    violations = np.flatnonzero(~condition)
    if not violations.size:
        return float(times[0]), True
    index = int(violations[-1] + 1)
    if index >= times.size:
        dt = float(times[-1] - times[-2]) if times.size > 1 else 0.0
        return float(times[-1] + dt), False
    return float(times[index]), True


def _overshoot_to_zero(values: np.ndarray) -> float:
    if not values.size or abs(float(values[0])) <= 1e-12:
        return 0.0
    if values[0] > 0.0:
        return max(0.0, -float(np.min(values)))
    return max(0.0, float(np.max(values)))


def _percent(values: np.ndarray) -> float:
    return float(100.0 * np.mean(values > 0.5)) if values.size else 0.0


def _tail_mask(times: np.ndarray, tail_window_s: float) -> np.ndarray:
    if not times.size:
        return np.array([], dtype=bool)
    return times >= max(float(times[0]), float(times[-1]) - tail_window_s)


def _row_values(rows: list[dict[str, Any]], key: str) -> np.ndarray:
    return np.asarray([float(row.get(key, 0.0)) for row in rows], dtype=float)


def compute_scenario_metrics(
    stage: str,
    result: LoiterRunResult,
    *,
    tail_window_s: float,
    baseline: dict[str, Any] | None = None,
) -> dict[str, Any]:
    rows = result.rows
    if not rows:
        return {
            "duration_s": 0.0,
            "pass": False,
            "rejected": True,
            "rejection_reasons": "missing scenario result",
            "crash_reason": "no rows",
        }

    times = _row_values(rows, "time")
    tail = _tail_mask(times, tail_window_s)
    rate_error_deg_s = np.rad2deg(_row_values(rows, "rate_error"))
    theta_deg = np.rad2deg(_row_values(rows, "theta"))
    omega_deg_s = np.rad2deg(_row_values(rows, "omega"))
    x_error = _row_values(rows, "x_error")
    x = _row_values(rows, "x")
    z = _row_values(rows, "z")
    vx = _row_values(rows, "vx")
    vane_cmd_deg = np.rad2deg(_row_values(rows, "vane_angle_cmd"))
    desired_moment = _row_values(rows, "desired_moment")
    rate_i = _row_values(rows, "rate_i")
    anti_windup = _row_values(rows, "anti_windup_correction")
    servo_angle_sat = _row_values(rows, "servo_angle_saturated")
    servo_rate_sat = _row_values(rows, "servo_rate_saturated")
    mixer_sat = _row_values(rows, "mixer_saturated")
    inhibition = _row_values(rows, "integrator_inhibited")
    moving_target = _row_values(rows, "moving_mass_target_m")
    moving_actual = _row_values(rows, "moving_mass_offset_m")
    moving_velocity = _row_values(rows, "moving_mass_velocity_m_s")
    moving_sat = _row_values(rows, "moving_mass_saturated")
    min_body_z = _row_values(rows, "min_body_z")

    if stage in {"rate_pd", "rate_i"}:
        settling_condition = np.abs(rate_error_deg_s) <= 2.0
        growth_signal = rate_error_deg_s
        growth_threshold = 5.0
    elif stage == "attitude_p":
        settling_condition = (np.abs(theta_deg) <= 1.0) & (np.abs(omega_deg_s) <= 3.0)
        growth_signal = theta_deg
        growth_threshold = 3.0
    else:
        settling_condition = (np.abs(x_error) <= 0.05) & (np.abs(vx) <= 0.05)
        growth_signal = x_error
        growth_threshold = 0.25
    settling_time_s, settled = _settling_time(times, settling_condition)

    growth_windows = np.array_split(growth_signal, 3)
    first_growth_rms, middle_growth_rms, tail_growth_rms = (
        _rms(window) for window in growth_windows
    )
    unbounded_growth = bool(
        tail_growth_rms > 2.0 * growth_threshold
        and middle_growth_rms > growth_threshold
        and middle_growth_rms > max(1e-9, 1.15 * first_growth_rms)
        and tail_growth_rms > 1.35 * middle_growth_rms
    )
    tail_saturation = max(
        _percent(servo_angle_sat[tail]),
        _percent(servo_rate_sat[tail]),
        _percent(mixer_sat[tail]),
    )
    excessive_sustained_saturation = bool(tail_saturation > 85.0)
    ground_contact = bool(np.any(min_body_z <= 0.0))
    attitude_limit_exceeded = bool(
        np.max(np.abs(theta_deg)) > float(result.scenario.max_theta_deg_limit)
    )

    metrics: dict[str, Any] = {
        "duration_s": float(times[-1]),
        "analytical_pass": bool(result.metrics.get("pass", False)),
        "rms_rate_error_deg_s": _rms(rate_error_deg_s),
        "tail_rms_rate_error_deg_s": _rms(rate_error_deg_s[tail]),
        "tail_mean_abs_rate_error_deg_s": float(np.mean(np.abs(rate_error_deg_s[tail]))),
        "rate_overshoot_deg_s": _overshoot_to_zero(rate_error_deg_s),
        "rate_zero_crossings": zero_crossing_count(rate_error_deg_s, 0.05),
        "rms_theta_deg": _rms(theta_deg),
        "tail_rms_theta_deg": _rms(theta_deg[tail]),
        "theta_overshoot_deg": _overshoot_to_zero(theta_deg),
        "tail_peak_to_peak_theta_deg": float(np.ptp(theta_deg[tail])) if np.any(tail) else 0.0,
        "max_omega_deg_s": float(np.max(np.abs(omega_deg_s))),
        "tail_rms_x_m": _rms(x_error[tail]),
        "tail_rms_vx_m_s": _rms(vx[tail]),
        "tail_peak_to_peak_x_m": float(np.ptp(x[tail])) if np.any(tail) else 0.0,
        "tail_path_length_m": path_length(x[tail], z[tail]),
        "x_error_zero_crossings": zero_crossing_count(x_error, 1e-5),
        "oscillation_peak_count": oscillation_peak_count(x_error[tail], 1e-4),
        "settling_time_s": settling_time_s,
        "settled": settled,
        "final_abs_x_error_m": abs(float(x_error[-1])),
        "rms_x_error_m": _rms(x_error),
        "vane_command_rms_deg": _rms(vane_cmd_deg),
        "vane_command_max_deg": float(np.max(np.abs(vane_cmd_deg))),
        "vane_angle_saturation_percent": _percent(servo_angle_sat),
        "servo_rate_saturation_percent": _percent(servo_rate_sat),
        "mixer_saturation_percent": _percent(mixer_sat),
        "control_effort_rms_Nm": _rms(desired_moment),
        "integrator_rms_Nm": _rms(rate_i),
        "integrator_max_abs_Nm": float(np.max(np.abs(rate_i))),
        "integrator_inhibition_percent": _percent(inhibition),
        "anti_windup_rms_Nm": _rms(anti_windup),
        "moving_mass_max_offset_m": float(np.max(np.abs(moving_actual))),
        "moving_mass_saturation_percent": _percent(moving_sat),
        "moving_mass_tracking_rms_m": _rms(moving_target - moving_actual),
        "moving_mass_target_reversals": direction_reversal_count(moving_target, 1e-8),
        "moving_mass_actual_reversals": zero_crossing_count(moving_velocity, 1e-8),
        "ground_contact": ground_contact,
        "attitude_limit_exceeded": attitude_limit_exceeded,
        "unbounded_growth": unbounded_growth,
        "excessive_sustained_saturation": excessive_sustained_saturation,
        "effective_parameter_mismatch": False,
    }

    if baseline is not None:
        baseline_rms_theta = max(_as_float(baseline.get("rms_theta_deg")), 1e-12)
        baseline_rms_x = max(_as_float(baseline.get("rms_x_error_m")), 1e-12)
        baseline_final_x = max(_as_float(baseline.get("final_abs_x_error_m")), 1e-12)
        baseline_tail_path = max(_as_float(baseline.get("tail_path_length_m")), 1e-12)
        baseline_vane_rms = max(_as_float(baseline.get("vane_command_rms_deg")), 1e-12)
        metrics.update(
            {
                "baseline_rms_theta_deg": baseline_rms_theta,
                "baseline_rms_x_error_m": baseline_rms_x,
                "baseline_final_abs_x_error_m": baseline_final_x,
                "baseline_tail_path_length_m": baseline_tail_path,
                "baseline_vane_command_rms_deg": baseline_vane_rms,
                "rms_theta_improvement_deg": baseline_rms_theta - metrics["rms_theta_deg"],
                "rms_theta_ratio": metrics["rms_theta_deg"] / baseline_rms_theta,
                "rms_x_change_m": metrics["rms_x_error_m"] - baseline_rms_x,
                "rms_x_ratio": metrics["rms_x_error_m"] / baseline_rms_x,
                "final_x_ratio": metrics["final_abs_x_error_m"] / baseline_final_x,
                "tail_path_ratio": metrics["tail_path_length_m"] / baseline_tail_path,
                "vane_command_rms_ratio": metrics["vane_command_rms_deg"] / baseline_vane_rms,
            }
        )

    crash_reason = result.crash_reason or str(result.metrics.get("crash_reason", ""))
    reasons: list[str] = []
    if crash_reason:
        reasons.append(f"crash: {crash_reason}")
    if ground_contact and "ground contact" not in crash_reason.lower():
        reasons.append("ground contact")
    if attitude_limit_exceeded:
        reasons.append("attitude safety limit exceeded")
    if unbounded_growth:
        reasons.append("unbounded growth")
    if excessive_sustained_saturation:
        reasons.append("excessive sustained saturation")
    numeric_values = [
        float(value)
        for key, value in metrics.items()
        if key not in {"settled", "ground_contact", "attitude_limit_exceeded", "unbounded_growth", "excessive_sustained_saturation", "effective_parameter_mismatch"}
        and isinstance(value, (int, float, np.integer, np.floating))
    ]
    if not all(math.isfinite(value) for value in numeric_values):
        reasons.append("NaN/Inf metric")
    if baseline is not None:
        materially_worse_hold = (
            metrics["rms_x_error_m"] > baseline_rms_x * 1.10 + 0.02
            or metrics["tail_path_length_m"] > baseline_tail_path * 1.15 + 0.02
            or metrics["final_abs_x_error_m"] > baseline_final_x * 1.15 + 0.02
        )
        if materially_worse_hold:
            reasons.append("materially worsened horizontal hold")

    metrics.update(
        {
            "pass": not reasons,
            "rejected": bool(reasons),
            "rejection_reasons": "; ".join(dict.fromkeys(reasons)),
            "crash_reason": crash_reason,
        }
    )
    return metrics


def normalized_score(row: dict[str, Any], stage: str) -> tuple[float, str]:
    components: dict[str, float] = {}
    weighted = 0.0
    total_weight = 0.0
    for spec in SCORE_SPECS[stage]:
        value = _as_float(row.get(spec.metric), float("nan"))
        if not math.isfinite(value):
            return float("inf"), json.dumps({spec.metric: "nonfinite"}, sort_keys=True)
        component = min(max(value / spec.reference_scale, 0.0), 5.0)
        components[spec.metric] = component
        weighted += component * spec.weight
        total_weight += spec.weight
    score = weighted / total_weight if total_weight else float("inf")
    return score, json.dumps(components, sort_keys=True, separators=(",", ":"))


def _symmetry_error(rows: Sequence[dict[str, Any]], stage: str) -> float:
    metric = "rms_theta_deg" if stage == "attitude_p" else "tail_rms_x_m"
    groups: dict[str, list[float]] = {}
    for row in rows:
        group = str(row.get("symmetry_group", ""))
        if group and _as_bool(row.get("primary_score", True)):
            groups.setdefault(group, []).append(_as_float(row.get(metric)))
    errors = []
    for values in groups.values():
        if len(values) == 2:
            errors.append(abs(values[0] - values[1]) / max(abs(values[0]), abs(values[1]), 1e-12))
    return float(np.mean(errors)) if errors else 0.0


def aggregate_candidates(
    stage: str,
    scenario_rows: Sequence[dict[str, Any]],
    expected_scenarios: Sequence[SearchScenario],
) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in scenario_rows:
        if row.get("stage") == stage:
            grouped.setdefault(str(row["candidate_key"]), []).append(dict(row))

    expected_names = {scenario.config.name for scenario in expected_scenarios}
    aggregates: list[dict[str, Any]] = []
    for candidate_key in sorted(grouped):
        rows = grouped[candidate_key]
        first = rows[0]
        present_names = {str(row["scenario_name"]) for row in rows}
        missing = sorted(expected_names - present_names)
        reasons = [str(row.get("rejection_reasons", "")) for row in rows if _as_bool(row.get("rejected"))]
        if missing:
            reasons.append("missing scenario result: " + ", ".join(missing))
        aggregate: dict[str, Any] = {
            "stage": stage,
            "candidate_id": first["candidate_id"],
            "candidate_key": candidate_key,
            "search_phase": first.get("search_phase", "full"),
            "scenario_count": len(rows),
            "primary_scenario_count": sum(_as_bool(row.get("primary_score", True)) for row in rows),
            "missing_scenario_count": len(missing),
            "rejected": bool(reasons),
            "rejection_reasons": "; ".join(dict.fromkeys(reason for reason in reasons if reason)),
        }
        for name in PARAMETER_FIELDS:
            aggregate[name] = first.get(name, "")
        primary_rows = [row for row in rows if _as_bool(row.get("primary_score", True))]
        for metric in AGGREGATE_METRIC_FIELDS:
            if metric == "left_right_symmetry_error":
                aggregate[metric] = _symmetry_error(primary_rows, stage)
                continue
            values = [_as_float(row.get(metric)) for row in primary_rows if row.get(metric, "") not in (None, "")]
            aggregate[metric] = float(np.mean(values)) if values else 0.0
        score, components = normalized_score(aggregate, stage)
        aggregate["normalized_score"] = score if not reasons else 1e9
        aggregate["score_components"] = components
        aggregate["rank"] = 0
        aggregate["selection_label"] = ""
        aggregates.append(aggregate)

    def tie_breaker(row: dict[str, Any]) -> tuple[Any, ...]:
        tail = _as_float(row.get("tail_path_length_m")) + _as_float(row.get("tail_rms_rate_error_deg_s")) + _as_float(row.get("tail_rms_theta_deg"))
        saturation = _as_float(row.get("vane_angle_saturation_percent")) + _as_float(row.get("servo_rate_saturation_percent"))
        effort = _as_float(row.get("control_effort_rms_Nm"))
        gain_size = sum(abs(_as_float(row.get(name))) for name in PARAMETER_FIELDS)
        score_bucket = round(_as_float(row.get("normalized_score")), 6)
        return (_as_bool(row.get("rejected")), score_bucket, tail, saturation, effort, gain_size, str(row["candidate_key"]))

    aggregates.sort(key=tie_breaker)
    for rank, row in enumerate(aggregates, 1):
        row["rank"] = rank
    mark_selection_roles(aggregates)
    return aggregates


def top_candidates(rows: Sequence[dict[str, Any]], count: int = 50) -> list[dict[str, Any]]:
    return [dict(row) for row in rows if not _as_bool(row.get("rejected"))][:count]


def mark_selection_roles(rows: list[dict[str, Any]]) -> None:
    accepted = [row for row in rows if not _as_bool(row.get("rejected"))]
    if not accepted:
        return
    role_keys = {
        "best aggregate": min(accepted, key=lambda row: (_as_float(row["normalized_score"]), str(row["candidate_key"]))),
        "most stable": min(
            accepted,
            key=lambda row: (
                _as_float(row.get("tail_path_length_m"))
                + _as_float(row.get("tail_rms_x_m"))
                + _as_float(row.get("tail_rms_rate_error_deg_s"))
                + _as_float(row.get("tail_rms_theta_deg")),
                str(row["candidate_key"]),
            ),
        ),
        "fastest settling": min(accepted, key=lambda row: (_as_float(row.get("settling_time_s")), str(row["candidate_key"]))),
        "lowest saturation": min(
            accepted,
            key=lambda row: (
                _as_float(row.get("vane_angle_saturation_percent"))
                + _as_float(row.get("servo_rate_saturation_percent"))
                + _as_float(row.get("mixer_saturation_percent")),
                str(row["candidate_key"]),
            ),
        ),
    }
    for role, selected in role_keys.items():
        current = str(selected.get("selection_label", ""))
        selected["selection_label"] = ", ".join(filter(None, (current, role)))


class ScenarioResultStore:
    def __init__(self, path: Path, *, resume: bool = True):
        self.path = path
        self.rows: list[dict[str, Any]] = []
        self.by_key: dict[str, dict[str, Any]] = {}
        if resume and path.exists():
            with path.open("r", newline="", encoding="utf-8-sig") as handle:
                for row in csv.DictReader(handle):
                    key = str(row.get("run_key", ""))
                    if not key:
                        raise ValueError("scenario result row is missing run_key")
                    if key in self.by_key:
                        raise ValueError(f"duplicate run key in resume CSV: {key}")
                    self.rows.append(row)
                    self.by_key[key] = row
        elif path.exists():
            path.unlink()

    def get(self, key: str) -> dict[str, Any] | None:
        return self.by_key.get(key)

    def add(self, row: dict[str, Any]) -> None:
        key = str(row["run_key"])
        if key in self.by_key:
            raise ValueError(f"duplicate run key: {key}")
        self.path.parent.mkdir(parents=True, exist_ok=True)
        write_header = not self.path.exists()
        with self.path.open("a", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=SCENARIO_CSV_FIELDS, extrasaction="ignore")
            if write_header:
                writer.writeheader()
            writer.writerow({field: row.get(field, "") for field in SCENARIO_CSV_FIELDS})
        self.rows.append(row)
        self.by_key[key] = row


def _controller_overrides(candidate: Candidate) -> dict[str, float]:
    return {
        name: float(candidate.parameters[name])
        for name in CONTROLLER_PARAMETER_FIELDS
        if name in candidate.parameters
    }


def _effective_mismatch(candidate: Candidate, result: LoiterRunResult) -> list[str]:
    mismatches = []
    for name, requested in _controller_overrides(candidate).items():
        effective_name = EFFECTIVE_FIELDS[name]
        effective = result.metrics.get(effective_name)
        if effective is None or not math.isclose(float(effective), requested, rel_tol=0.0, abs_tol=1e-12):
            mismatches.append(f"{name}: requested={requested:g}, effective={effective!r}")
    return mismatches


def _scenario_run_key(
    candidate: Candidate,
    scenario: SearchScenario,
    param_path: str | Path,
    tail_window_s: float,
) -> str:
    config_json = json.dumps(asdict(scenario.config), sort_keys=True, separators=(",", ":"))
    config_digest = hashlib.sha1(config_json.encode("utf-8")).hexdigest()[:10]
    source_digest = hashlib.sha1(Path(param_path).read_bytes()).hexdigest()[:10]
    return (
        f"{candidate.key}|scenario={scenario.config.name}|cfg={config_digest}"
        f"|tail={_format_number(tail_window_s)}|source={source_digest}"
    )


def _scenario_row(
    stage: str,
    candidate: Candidate,
    scenario: SearchScenario,
    param_path: str | Path,
    result: LoiterRunResult,
    *,
    tail_window_s: float,
    baseline: dict[str, Any] | None = None,
    run_key: str | None = None,
    enforce_primary_pass: bool = False,
) -> dict[str, Any]:
    metrics = compute_scenario_metrics(stage, result, tail_window_s=tail_window_s, baseline=baseline)
    mismatches = _effective_mismatch(candidate, result)
    if mismatches:
        metrics["effective_parameter_mismatch"] = True
        metrics["rejected"] = True
        metrics["pass"] = False
        reasons = str(metrics.get("rejection_reasons", ""))
        mismatch_reason = "effective parameter mismatch: " + "; ".join(mismatches)
        metrics["rejection_reasons"] = "; ".join(filter(None, (reasons, mismatch_reason)))
    if (
        enforce_primary_pass
        and stage == "loiter_xy"
        and scenario.primary_score
        and not _as_bool(metrics.get("analytical_pass"))
    ):
        reasons = str(metrics.get("rejection_reasons", ""))
        metrics["rejected"] = True
        metrics["pass"] = False
        metrics["rejection_reasons"] = "; ".join(
            filter(None, (reasons, "primary scenario analytical recovery threshold failed"))
        )
    row: dict[str, Any] = {
        "run_key": run_key or _scenario_run_key(candidate, scenario, param_path, tail_window_s),
        "stage": stage,
        "candidate_id": candidate.candidate_id,
        "candidate_key": candidate.key,
        "search_phase": candidate.search_phase,
        "scenario_name": scenario.config.name,
        "symmetry_group": scenario.symmetry_group,
        "primary_score": scenario.primary_score,
        "param_file": str(param_path),
        **candidate.parameters,
        **metrics,
    }
    for name, effective_name in EFFECTIVE_FIELDS.items():
        row[effective_name] = result.metrics.get(effective_name, "")
    return row


def _run_candidate_scenarios(
    stage: str,
    candidates: Sequence[Candidate],
    scenarios: Sequence[SearchScenario],
    *,
    param_path: str | Path,
    store: ScenarioResultStore,
    tail_window_s: float,
    baseline_by_scenario: dict[str, dict[str, Any]] | None = None,
    moving_mass_geometry: bool = False,
    quick: bool = False,
) -> list[dict[str, Any]]:
    total = len(candidates) * len(scenarios)
    completed = 0
    started = time.perf_counter()
    stage_rows: list[dict[str, Any]] = []
    for candidate_index, candidate in enumerate(candidates, 1):
        for scenario in scenarios:
            run_key = _scenario_run_key(candidate, scenario, param_path, tail_window_s)
            cached = store.get(run_key)
            if cached is not None:
                row = cached
            else:
                scenario_cfg = scenario.config
                rb_overrides: dict[str, Any] = {
                    "moving_mass": {
                        "enabled": bool(moving_mass_geometry),
                        "use_total_com_geometry": bool(moving_mass_geometry),
                        "use_legacy_gravity_offset_moment": not bool(moving_mass_geometry),
                    }
                }
                if moving_mass_geometry:
                    scenario_cfg = replace(
                        scenario_cfg,
                        moving_mass_enabled=True,
                        moving_mass_target_m=0.0,
                        moving_mass_assist_gain_m_per_Nm=float(
                            candidate.parameters["moving_mass_assist_gain_m_per_Nm"]
                        ),
                    )
                result = run_headless_loiter(
                    param_path,
                    scenario_cfg,
                    rb_overrides=rb_overrides,
                    controller_overrides=_controller_overrides(candidate),
                )
                baseline = (baseline_by_scenario or {}).get(scenario.config.name)
                row = _scenario_row(
                    stage,
                    candidate,
                    scenario,
                    param_path,
                    result,
                    tail_window_s=tail_window_s,
                    baseline=baseline,
                    run_key=run_key,
                    enforce_primary_pass=not quick,
                )
                store.add(row)
            stage_rows.append(row)
            completed += 1
        elapsed = time.perf_counter() - started
        if quick or candidate_index == 1 or candidate_index == len(candidates) or candidate_index % 10 == 0:
            rate = completed / max(elapsed, 1e-9)
            remaining = (total - completed) / max(rate, 1e-9)
            print(
                f"[{stage}] candidate {candidate_index}/{len(candidates)} "
                f"({completed}/{total} scenario runs, ETA {remaining:.1f}s)"
            )
    return stage_rows


def _run_moving_mass_baselines(
    controller_parameters: dict[str, float],
    scenarios: Sequence[SearchScenario],
    *,
    param_path: str | Path,
    store: ScenarioResultStore,
    tail_window_s: float,
) -> dict[str, dict[str, Any]]:
    candidate = Candidate("moving_mass_baseline", controller_parameters, "baseline")
    baselines: dict[str, dict[str, Any]] = {}
    for scenario in scenarios:
        run_key = _scenario_run_key(candidate, scenario, param_path, tail_window_s)
        cached = store.get(run_key)
        if cached is not None:
            baselines[scenario.config.name] = cached
            continue
        scenario_cfg = replace(
            scenario.config,
            moving_mass_enabled=False,
            moving_mass_target_m=0.0,
            moving_mass_assist_gain_m_per_Nm=0.0,
        )
        result = run_headless_loiter(
            param_path,
            scenario_cfg,
            rb_overrides={
                "moving_mass": {
                    "enabled": False,
                    "use_total_com_geometry": True,
                    "use_legacy_gravity_offset_moment": False,
                }
            },
            controller_overrides=controller_parameters,
        )
        row = _scenario_row(
            "moving_mass_baseline",
            candidate,
            scenario,
            param_path,
            result,
            tail_window_s=tail_window_s,
            run_key=run_key,
        )
        store.add(row)
        baselines[scenario.config.name] = row
    return baselines


def write_csv_rows(path: Path, rows: Sequence[dict[str, Any]], fields: Sequence[str]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})
    return path


def read_csv_rows(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def best_aggregate_row(rows: Sequence[dict[str, Any]]) -> dict[str, Any] | None:
    return next((dict(row) for row in rows if not _as_bool(row.get("rejected"))), None)


def _parameters_from_row(row: dict[str, Any] | None, names: Sequence[str], defaults: dict[str, float]) -> dict[str, float]:
    return {
        name: _as_float(row.get(name), defaults[name]) if row is not None else defaults[name]
        for name in names
    }


def best_parameter_rows(stage_aggregates: dict[str, list[dict[str, Any]]]) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    roles = ("best aggregate", "most stable", "fastest settling", "lowest saturation")
    for stage in STAGES:
        rows = stage_aggregates.get(stage, [])
        for role in roles:
            selected = next(
                (row for row in rows if role in str(row.get("selection_label", "")).split(", ")),
                None,
            )
            if selected is None:
                continue
            output.append(
                {
                    "row_key": f"{stage}|{role}",
                    "stage": stage,
                    "role": role,
                    "candidate_id": selected["candidate_id"],
                    "rank": selected["rank"],
                    "normalized_score": selected["normalized_score"],
                    **{name: selected.get(name, "") for name in PARAMETER_FIELDS},
                    "rejected": selected["rejected"],
                }
            )
    return output


def _excel_value(value: Any) -> Any:
    if isinstance(value, (np.integer, np.floating)):
        return value.item()
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, sort_keys=True)
    return value


def write_workbook(
    path: Path,
    stage_aggregates: dict[str, list[dict[str, Any]]],
    scenario_rows: Sequence[dict[str, Any]],
    metadata_rows: Sequence[dict[str, Any]],
) -> Path:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    workbook.remove(workbook.active)
    sheet_payloads: list[tuple[str, Sequence[dict[str, Any]], Sequence[str]]] = [
        ("01_rate_pd_all", stage_aggregates.get("rate_pd", []), AGGREGATE_CSV_FIELDS),
        ("02_rate_pd_top50", top_candidates(stage_aggregates.get("rate_pd", []), 50), AGGREGATE_CSV_FIELDS),
        ("03_rate_i_all", stage_aggregates.get("rate_i", []), AGGREGATE_CSV_FIELDS),
        ("04_attitude_p_all", stage_aggregates.get("attitude_p", []), AGGREGATE_CSV_FIELDS),
        ("05_loiter_xy_all", stage_aggregates.get("loiter_xy", []), AGGREGATE_CSV_FIELDS),
        ("06_loiter_xy_top50", top_candidates(stage_aggregates.get("loiter_xy", []), 50), AGGREGATE_CSV_FIELDS),
        ("07_moving_mass_gain_all", stage_aggregates.get("moving_mass_gain", []), AGGREGATE_CSV_FIELDS),
        (
            "08_scenario_summary",
            sorted(scenario_rows, key=lambda row: (str(row.get("stage")), str(row.get("candidate_key")), str(row.get("scenario_name")))),
            SCENARIO_CSV_FIELDS,
        ),
        (
            "09_best_parameters",
            best_parameter_rows(stage_aggregates),
            ("row_key", "stage", "role", "candidate_id", "rank", "normalized_score", *PARAMETER_FIELDS, "rejected"),
        ),
        ("10_metadata", metadata_rows, ("key", "value", "unit", "notes")),
    ]

    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(color="FFFFFF", bold=True)
    top_fill = PatternFill("solid", fgColor="E2F0D9")
    rejected_fill = PatternFill("solid", fgColor="FCE4D6")
    thin_gray = Side(style="thin", color="D9E2F3")

    for sheet_index, (sheet_name, rows, fields) in enumerate(sheet_payloads, 1):
        sheet = workbook.create_sheet(sheet_name)
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A2"
        for col_index, field in enumerate(fields, 1):
            cell = sheet.cell(row=1, column=col_index, value=field)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin_gray)
        for row_index, row in enumerate(rows, 2):
            for col_index, field in enumerate(fields, 1):
                cell = sheet.cell(row=row_index, column=col_index, value=_excel_value(row.get(field, "")))
                cell.alignment = Alignment(vertical="top", wrap_text=field in {"rejection_reasons", "notes", "value"})
                if field.endswith("percent"):
                    cell.number_format = "0.00"
                elif field in {"rank", "scenario_count", "primary_scenario_count", "missing_scenario_count"} or field.endswith("crossings") or field.endswith("reversals") or field.endswith("peak_count"):
                    cell.number_format = "0"
                elif isinstance(cell.value, float):
                    cell.number_format = "0.000000"
        max_row = max(1, len(rows) + 1)
        max_col = len(fields)
        sheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
        if rows:
            table = Table(displayName=f"GridTable{sheet_index:02d}", ref=f"A1:{get_column_letter(max_col)}{max_row}")
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
            sheet.add_table(table)
        for col_index, field in enumerate(fields, 1):
            samples = [str(field)] + [str(row.get(field, "")) for row in rows[:100]]
            width = min(34, max(10, max(len(sample) for sample in samples) + 2))
            if field in {"candidate_key", "run_key", "rejection_reasons", "notes", "value"}:
                width = min(48, max(width, 24))
            sheet.column_dimensions[get_column_letter(col_index)].width = width
        if rows:
            field_index = {field: index + 1 for index, field in enumerate(fields)}
            if "normalized_score" in field_index:
                column = get_column_letter(field_index["normalized_score"])
                sheet.conditional_formatting.add(
                    f"{column}2:{column}{max_row}",
                    ColorScaleRule(start_type="min", start_color="63BE7B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="F8696B"),
                )
            if "selection_label" in field_index:
                column = get_column_letter(field_index["selection_label"])
                sheet.conditional_formatting.add(
                    f"A2:{get_column_letter(max_col)}{max_row}",
                    FormulaRule(formula=[f'LEN(${column}2)>0'], fill=top_fill),
                )
            if "rejected" in field_index:
                column = get_column_letter(field_index["rejected"])
                sheet.conditional_formatting.add(
                    f"A2:{get_column_letter(max_col)}{max_row}",
                    FormulaRule(formula=[f'OR(${column}2=TRUE,${column}2="True",${column}2="1")'], fill=rejected_fill),
                )
        sheet.row_dimensions[1].height = 28

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def write_profiles(
    stage_aggregates: dict[str, list[dict[str, Any]]],
    *,
    vane_source: Path,
    moving_mass_source: Path,
    output_directory: Path,
) -> tuple[Path, Path]:
    with vane_source.open("r", encoding="utf-8-sig") as handle:
        vane_profile = json.load(handle)
    with moving_mass_source.open("r", encoding="utf-8-sig") as handle:
        moving_profile = json.load(handle)

    rate_pd = best_aggregate_row(stage_aggregates.get("rate_pd", []))
    rate_i = best_aggregate_row(stage_aggregates.get("rate_i", []))
    attitude = best_aggregate_row(stage_aggregates.get("attitude_p", []))
    loiter = best_aggregate_row(stage_aggregates.get("loiter_xy", []))
    moving = best_aggregate_row(stage_aggregates.get("moving_mass_gain", []))
    selected_rows = [row for row in (rate_pd, rate_i, attitude, loiter) if row is not None]
    selected: dict[str, float] = {}
    for row in selected_rows:
        for name in CONTROLLER_PARAMETER_FIELDS:
            if row.get(name, "") not in (None, ""):
                selected[name] = _as_float(row[name])

    for profile in (vane_profile, moving_profile):
        controller = profile.setdefault("controller", {})
        for name, value in selected.items():
            controller[name] = value
    moving_profile.setdefault("analysis", {})["moving_mass_assist_gain_m_per_Nm"] = (
        _as_float(moving.get("moving_mass_assist_gain_m_per_Nm")) if moving is not None else 0.0
    )
    moving_profile["analysis"]["note"] = "Analysis-only proportional assist gain; not a PID field."

    output_directory.mkdir(parents=True, exist_ok=True)
    vane_path = output_directory / "loiter_tuned_vane_only.json"
    moving_path = output_directory / "moving_mass_prototype_2kg_tuned.json"
    vane_path.write_text(json.dumps(vane_profile, indent=2) + "\n", encoding="utf-8")
    moving_path.write_text(json.dumps(moving_profile, indent=2) + "\n", encoding="utf-8")
    return vane_path, moving_path


def _ribbon_assessment(loiter_rows: Sequence[dict[str, Any]]) -> str:
    best = best_aggregate_row(loiter_rows)
    if best is None:
        return "not assessed because no non-rejected LOITER candidate was available"
    eliminated = (
        _as_float(best.get("tail_rms_x_m")) <= 0.02
        and _as_float(best.get("tail_peak_to_peak_x_m")) <= 0.05
        and _as_float(best.get("tail_path_length_m")) <= 0.05
        and _as_float(best.get("tail_rms_vx_m_s")) <= 0.02
    )
    if eliminated:
        return "eliminated under the deterministic primary scenarios and documented tail thresholds"
    baseline = next(
        (
            row
            for row in loiter_rows
            if math.isclose(_as_float(row.get("psc_ne_pos_p")), 0.8, abs_tol=1e-12)
            and math.isclose(_as_float(row.get("psc_ne_vel_p")), 1.1, abs_tol=1e-12)
            and not _as_bool(row.get("rejected"))
        ),
        None,
    )
    if baseline is not None and _as_float(best.get("normalized_score")) < _as_float(baseline.get("normalized_score")):
        return "reduced but not eliminated under the documented tail thresholds"
    return "not eliminated and not measurably reduced versus the mode-matched 0.8/1.1 baseline"


def write_markdown_summary(
    path: Path,
    stage_aggregates: dict[str, list[dict[str, Any]]],
    metadata: dict[str, Any],
) -> Path:
    lines = [
        "# Controller Gain Grid Search",
        "",
        "This report ranks deterministic analytical simulations. It does not claim real-flight optimality.",
        "Raw metrics are normalized by the documented reference scales before weighting; raw values with different units are never summed directly.",
        "",
        "## Reproduction",
        "",
        "```powershell",
        ".venv\\Scripts\\python.exe sweep_controller_gains.py --stage all --output-dir results/analysis/controller_grid_search",
        ".venv\\Scripts\\python.exe sweep_controller_gains.py --stage all --quick --output-dir results/analysis/controller_grid_search_quick",
        "```",
        "",
        f"- Git SHA: `{metadata.get('git_sha', '')}`",
        f"- Parameter source: `{metadata.get('vane_param_source', '')}`",
        f"- Tail window: `{metadata.get('tail_window_s', '')}` s",
        f"- Runtime: `{metadata.get('runtime_s', '')}` s",
        f"- Ribbon/comet assessment: **{_ribbon_assessment(stage_aggregates.get('loiter_xy', []))}**.",
        "- `psc_ne_vel_i` and `psc_ne_vel_d` remain inactive and were not swept or optimized.",
        "",
        "## Best Aggregate Candidates",
        "",
        "| stage | candidate | score | selected parameters | rejected |",
        "| --- | --- | ---: | --- | --- |",
    ]
    for stage in STAGES:
        row = best_aggregate_row(stage_aggregates.get(stage, []))
        if row is None:
            lines.append(f"| {stage} | n/a | n/a | no non-rejected candidate | yes |")
            continue
        params = ", ".join(
            f"{name}={_format_number(_as_float(row[name]))}"
            for name in PARAMETER_FIELDS
            if row.get(name, "") not in (None, "")
        )
        lines.append(
            f"| {stage} | {row['candidate_id']} | {_as_float(row['normalized_score']):.6f} | {params} | {row['rejected']} |"
        )
    lines += ["", "## Grid And Run Counts", "", "| stage | candidates | scenario rows |", "| --- | ---: | ---: |"]
    for stage in STAGES:
        rows = stage_aggregates.get(stage, [])
        lines.append(f"| {stage} | {len(rows)} | {sum(int(_as_float(row.get('scenario_count'))) for row in rows)} |")
    lines += [
        "",
        "## Rejection And Tie-Break Rules",
        "",
        "Candidates are rejected for crashes, ground contact, non-finite data, attitude-limit violations, unbounded growth, excessive sustained saturation, missing scenarios, duplicate run keys, or effective-parameter mismatches. Moving-mass candidates are also rejected when horizontal hold is materially worse than the total-COM centered baseline.",
        "Ties favor lower tail oscillation, then lower saturation, lower control effort, and finally smaller gain magnitude.",
        "Authority-stress LOITER rows participate in hard rejection and robustness reporting but not the primary aggregate score.",
        "",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")
    return path


STAGE_CSV_FILES = {
    "rate_pd": "01_rate_pd_all.csv",
    "rate_i": "03_rate_i_all.csv",
    "attitude_p": "04_attitude_p_all.csv",
    "loiter_xy": "05_loiter_xy_all.csv",
    "moving_mass_gain": "07_moving_mass_gain_all.csv",
}


@dataclass(frozen=True)
class WorkflowOptions:
    stage: str = "all"
    output_dir: Path = Path("results/analysis/controller_grid_search")
    quick: bool = False
    resume: bool = True
    tail_window_s: float = 2.0
    vane_param_source: Path = Path("params/loiter_example.json")
    moving_mass_param_source: Path = Path("params/moving_mass_prototype_2kg.json")
    profile_output_dir: Path = Path("params")
    top_pd_count: int = 3


def _git_sha() -> str:
    try:
        return subprocess.check_output(
            ["git", "rev-parse", "HEAD"], text=True, encoding="utf-8"
        ).strip()
    except (OSError, subprocess.CalledProcessError):
        return "unknown"


def _metadata_rows(metadata: dict[str, Any]) -> list[dict[str, Any]]:
    units = {
        "tail_window_s": "s",
        "runtime_s": "s",
        "scenario_run_count": "runs",
        "candidate_count": "candidates",
    }
    notes = {
        "score_specs": "Dimensionless score components use metric/reference_scale before weighting.",
        "inactive_parameters": "Existing fields are inactive in LOITER and were not swept or optimized.",
        "moving_mass_gain": "Analysis proportional gain; it is not a PID term.",
    }
    return [
        {
            "key": key,
            "value": value if isinstance(value, (str, int, float, bool)) else json.dumps(value, sort_keys=True),
            "unit": units.get(key, ""),
            "notes": notes.get(key, ""),
        }
        for key, value in metadata.items()
    ]


def _write_stage_exports(
    output_dir: Path,
    stage_aggregates: dict[str, list[dict[str, Any]]],
) -> None:
    for stage, filename in STAGE_CSV_FILES.items():
        write_csv_rows(output_dir / filename, stage_aggregates.get(stage, []), AGGREGATE_CSV_FIELDS)
    write_csv_rows(
        output_dir / "02_rate_pd_top50.csv",
        top_candidates(stage_aggregates.get("rate_pd", []), 50),
        AGGREGATE_CSV_FIELDS,
    )
    write_csv_rows(
        output_dir / "06_loiter_xy_top50.csv",
        top_candidates(stage_aggregates.get("loiter_xy", []), 50),
        AGGREGATE_CSV_FIELDS,
    )
    write_csv_rows(
        output_dir / "09_best_parameters.csv",
        best_parameter_rows(stage_aggregates),
        ("row_key", "stage", "role", "candidate_id", "rank", "normalized_score", *PARAMETER_FIELDS, "rejected"),
    )


def _load_stage_aggregates(output_dir: Path) -> dict[str, list[dict[str, Any]]]:
    return {
        stage: read_csv_rows(output_dir / filename)
        for stage, filename in STAGE_CSV_FILES.items()
    }


def run_workflow(options: WorkflowOptions) -> dict[str, Any]:
    if options.stage != "all" and options.stage not in STAGES:
        raise ValueError(f"unknown stage {options.stage!r}")
    if options.tail_window_s <= 0.0:
        raise ValueError("tail_window_s must be positive")
    if options.top_pd_count <= 0:
        raise ValueError("top_pd_count must be positive")

    output_dir = Path(options.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    store = ScenarioResultStore(output_dir / "scenario_results.csv", resume=options.resume)
    aggregates = _load_stage_aggregates(output_dir)
    selected_stages = set(STAGES if options.stage == "all" else (options.stage,))
    started = time.perf_counter()

    _rb, _ui, controller_cfg = load_interactive_config(options.vane_param_source)
    defaults = {name: float(getattr(controller_cfg, name)) for name in CONTROLLER_PARAMETER_FIELDS}

    if "rate_pd" in selected_stages:
        scenarios = rate_pd_scenarios(options.quick)
        coarse = rate_pd_coarse_candidates(options.quick)
        coarse_rows = _run_candidate_scenarios(
            "rate_pd", coarse, scenarios,
            param_path=options.vane_param_source, store=store,
            tail_window_s=options.tail_window_s, quick=options.quick,
        )
        coarse_aggregates = aggregate_candidates("rate_pd", coarse_rows, scenarios)
        fine_seed_count = 2 if options.quick else 5
        fine = rate_pd_fine_candidates(
            top_candidates(coarse_aggregates, fine_seed_count),
            existing_keys={candidate.key for candidate in coarse},
        )
        fine_rows = _run_candidate_scenarios(
            "rate_pd", fine, scenarios,
            param_path=options.vane_param_source, store=store,
            tail_window_s=options.tail_window_s, quick=options.quick,
        )
        aggregates["rate_pd"] = aggregate_candidates("rate_pd", [*coarse_rows, *fine_rows], scenarios)

    pd_rows = top_candidates(aggregates.get("rate_pd", []), 2 if options.quick else options.top_pd_count)
    if not pd_rows:
        pd_rows = [
            {
                "atc_rat_pit_p": defaults["atc_rat_pit_p"],
                "atc_rat_pit_i": 0.0,
                "atc_rat_pit_d": defaults["atc_rat_pit_d"],
            }
        ]
    if "rate_i" in selected_stages:
        scenarios = rate_i_scenarios(options.quick)
        candidates = rate_i_candidates(pd_rows, options.quick)
        rows = _run_candidate_scenarios(
            "rate_i", candidates, scenarios,
            param_path=options.vane_param_source, store=store,
            tail_window_s=options.tail_window_s, quick=options.quick,
        )
        aggregates["rate_i"] = aggregate_candidates("rate_i", rows, scenarios)

    best_rate_i = best_aggregate_row(aggregates.get("rate_i", []))
    if best_rate_i is None:
        best_rate_i = pd_rows[0]
    rate_parameters = _parameters_from_row(
        best_rate_i,
        ("atc_rat_pit_p", "atc_rat_pit_i", "atc_rat_pit_d"),
        defaults,
    )
    if "attitude_p" in selected_stages:
        scenarios = attitude_p_scenarios(options.quick)
        candidates = attitude_p_candidates(rate_parameters, options.quick)
        rows = _run_candidate_scenarios(
            "attitude_p", candidates, scenarios,
            param_path=options.vane_param_source, store=store,
            tail_window_s=options.tail_window_s, quick=options.quick,
        )
        aggregates["attitude_p"] = aggregate_candidates("attitude_p", rows, scenarios)

    best_attitude = best_aggregate_row(aggregates.get("attitude_p", []))
    attitude_parameters = {
        **rate_parameters,
        "atc_ang_pit_p": _as_float(
            best_attitude.get("atc_ang_pit_p") if best_attitude else None,
            defaults["atc_ang_pit_p"],
        ),
    }
    if "loiter_xy" in selected_stages:
        scenarios = loiter_xy_scenarios(options.quick)
        candidates = loiter_xy_candidates(attitude_parameters, options.quick)
        rows = _run_candidate_scenarios(
            "loiter_xy", candidates, scenarios,
            param_path=options.vane_param_source, store=store,
            tail_window_s=options.tail_window_s, quick=options.quick,
        )
        aggregates["loiter_xy"] = aggregate_candidates("loiter_xy", rows, scenarios)

    best_loiter = best_aggregate_row(aggregates.get("loiter_xy", []))
    loiter_parameters = {
        **attitude_parameters,
        "psc_ne_pos_p": _as_float(
            best_loiter.get("psc_ne_pos_p") if best_loiter else None,
            defaults["psc_ne_pos_p"],
        ),
        "psc_ne_vel_p": _as_float(
            best_loiter.get("psc_ne_vel_p") if best_loiter else None,
            defaults["psc_ne_vel_p"],
        ),
    }
    if "moving_mass_gain" in selected_stages:
        scenarios = moving_mass_gain_scenarios(options.quick)
        baseline_by_scenario = _run_moving_mass_baselines(
            loiter_parameters,
            scenarios,
            param_path=options.moving_mass_param_source,
            store=store,
            tail_window_s=options.tail_window_s,
        )
        candidates = moving_mass_gain_candidates(loiter_parameters, options.quick)
        rows = _run_candidate_scenarios(
            "moving_mass_gain", candidates, scenarios,
            param_path=options.moving_mass_param_source, store=store,
            tail_window_s=options.tail_window_s,
            baseline_by_scenario=baseline_by_scenario,
            moving_mass_geometry=True,
            quick=options.quick,
        )
        aggregates["moving_mass_gain"] = aggregate_candidates("moving_mass_gain", rows, scenarios)

    runtime_s = time.perf_counter() - started
    _write_stage_exports(output_dir, aggregates)
    vane_profile, moving_profile = write_profiles(
        aggregates,
        vane_source=Path(options.vane_param_source),
        moving_mass_source=Path(options.moving_mass_param_source),
        output_directory=Path(options.profile_output_dir),
    )
    metadata: dict[str, Any] = {
        "git_sha": _git_sha(),
        "quick_mode": options.quick,
        "selected_stage": options.stage,
        "vane_param_source": str(options.vane_param_source),
        "moving_mass_param_source": str(options.moving_mass_param_source),
        "tail_window_s": options.tail_window_s,
        "runtime_s": round(runtime_s, 6),
        "candidate_count": sum(len(rows) for rows in aggregates.values()),
        "scenario_run_count": len(store.rows),
        "stage_scenarios": {
            "rate_pd": [scenario.config.name for scenario in rate_pd_scenarios(options.quick)],
            "rate_i": [scenario.config.name for scenario in rate_i_scenarios(options.quick)],
            "attitude_p": [scenario.config.name for scenario in attitude_p_scenarios(options.quick)],
            "loiter_xy": [scenario.config.name for scenario in loiter_xy_scenarios(options.quick)],
            "moving_mass_gain": [scenario.config.name for scenario in moving_mass_gain_scenarios(options.quick)],
        },
        "score_specs": {
            stage: [asdict(spec) for spec in specs]
            for stage, specs in SCORE_SPECS.items()
        },
        "grid_ranges": {
            "rate_pd": "P 0.005..0.080 step 0.005; D 0.0000..0.0080 step 0.0005; local half-step fine search",
            "rate_i": "I 0.000..0.030 step 0.002",
            "attitude_p": "Angle P 2.0..10.0 step 0.5",
            "loiter_xy": "position P 0.1..1.5 step 0.1; velocity P 0.2..2.5 step 0.1",
            "moving_mass_gain": "0.000..0.080 step 0.0025",
        },
        "inactive_parameters": "psc_ne_vel_i, psc_ne_vel_d",
        "moving_mass_gain": "moving_mass_assist_gain_m_per_Nm * desired_pitch_moment",
        "vane_profile": str(vane_profile),
        "moving_mass_profile": str(moving_profile),
    }
    metadata_rows = _metadata_rows(metadata)
    write_csv_rows(output_dir / "10_metadata.csv", metadata_rows, ("key", "value", "unit", "notes"))
    workbook_path = write_workbook(
        output_dir / "controller_gain_search.xlsx",
        aggregates,
        store.rows,
        metadata_rows,
    )
    markdown_path = write_markdown_summary(
        output_dir / "controller_gain_search_summary.md",
        aggregates,
        metadata,
    )
    return {
        "stage_aggregates": aggregates,
        "scenario_rows": store.rows,
        "metadata": metadata,
        "runtime_s": runtime_s,
        "workbook_path": workbook_path,
        "markdown_path": markdown_path,
        "vane_profile_path": vane_profile,
        "moving_mass_profile_path": moving_profile,
        "ribbon_assessment": _ribbon_assessment(aggregates.get("loiter_xy", [])),
    }
